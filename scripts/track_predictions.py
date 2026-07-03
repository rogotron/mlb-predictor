"""Log model predictions to Excel and fill in actual scores.

Each day, run --log to append today's predictions.
Run --actuals to auto-fill the previous day's real scores from the schedule cache.
Run --actuals --all-missing to fill every pending final score in the tracker.

Usage:
    python scripts/track_predictions.py --log
    python scripts/track_predictions.py --actuals
    python scripts/track_predictions.py --actuals --date 2026-05-06
    python scripts/track_predictions.py --actuals --all-missing
    python scripts/track_predictions.py --log --actuals
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import openpyxl
import requests
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.data.update import today_in_schedule_timezone
from src.utils.logging import configure_logging

TRACKER_PATH = Path("mlb_predictions_tracker.xlsx")
CHECK_MARK = "\u2713"
X_MARK = "\u2717"

HEADERS = [
    "Date",
    "Away",
    "Home",
    "Away SP",
    "Home SP",
    "Predicted Winner",
    "Away Pred",
    "Home Pred",
    "Win Prob %",
    "Confidence",
    "Away Actual",
    "Home Actual",
    "Actual Winner",
    "Correct?",
    "Game PK",
]

COL_AWAY = 2
COL_HOME = 3
COL_PRED_WINNER = 6
COL_AWAY_PRED = 7
COL_HOME_PRED = 8
COL_WIN_PROB = 9
COL_AWAY_ACT = 11
COL_HOME_ACT = 12
COL_ACT_WINNER = 13
COL_CORRECT = 14
COL_GAME_PK = 15

INK = "1C1410"
PARCH = "EDE3CC"
CARD = "DDD0B8"
RED = "9A1515"
GREEN = "1E4D1E"
GOLD = "7A5A1A"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = INK, size: int = 10) -> Font:
    return Font(name="Courier New", bold=bold, color=color, size=size)


def _border() -> Border:
    thin = Side(style="thin", color="8A7A65")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _load_or_create() -> openpyxl.Workbook:
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
        _ensure_predictions_sheet(wb["Predictions"])
        if "Summary" in wb.sheetnames:
            _ensure_summary_sheet(wb["Summary"])
        else:
            _init_summary_sheet(wb.create_sheet("Summary"))
        return wb

    wb = openpyxl.Workbook()
    _init_predictions_sheet(wb.active)
    _init_summary_sheet(wb.create_sheet("Summary"))
    return wb


def _init_predictions_sheet(ws) -> None:
    ws.title = "Predictions"
    _ensure_predictions_sheet(ws)


def _ensure_predictions_sheet(ws) -> None:
    ws.freeze_panes = "A2"

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Courier New", bold=True, color=PARCH, size=10)
        cell.fill = _fill(INK)
        cell.alignment = _center()
        cell.border = _border()

    widths = [12, 22, 22, 20, 20, 22, 11, 11, 11, 13, 12, 12, 22, 10, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.column_dimensions[get_column_letter(COL_GAME_PK)].hidden = True
    ws.row_dimensions[1].height = 20

    for row_num in range(2, ws.max_row + 1):
        if ws.cell(row_num, 1).value:
            _ensure_result_formulas(ws, row_num)

    _backfill_game_pks_from_static(ws)
    _apply_conditional_formatting(ws)


def _init_summary_sheet(ws) -> None:
    ws.title = "Summary"
    _ensure_summary_sheet(ws)


def _ensure_summary_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    def row(r: int, label: str, formula: str) -> None:
        ws.cell(r, 1, label).font = _font(bold=True)
        ws.cell(r, 2, formula).font = _font()
        ws.cell(r, 2).alignment = _center()

    ws.cell(1, 1, "MLB PREDICTIONS TRACKER").font = Font(
        name="Courier New", bold=True, color=RED, size=13
    )
    row(3, "Total Games Logged", "=COUNTA(Predictions!A2:A5000)")
    row(4, "Actuals Entered", "=COUNTA(Predictions!K2:K5000)")
    row(5, "Correct Predictions", f'=COUNTIF(Predictions!N2:N5000,"{CHECK_MARK}")')
    row(6, "Incorrect", f'=COUNTIF(Predictions!N2:N5000,"{X_MARK}")')
    row(7, "Accuracy %", '=IFERROR(B5/B4,"-")')
    ws.cell(7, 2).number_format = "0.0%"

    ws.cell(9, 1, "BY CONFIDENCE").font = _font(bold=True, color=GOLD)
    for i, conf in enumerate(["HIGH", "MODERATE", "LOW", "MARGINAL"], 10):
        ws.cell(i, 1, conf).font = _font()
        ws.cell(
            i,
            2,
            f'=IFERROR(COUNTIFS(Predictions!J2:J5000,"{conf}",Predictions!N2:N5000,"{CHECK_MARK}")'
            f'/COUNTIFS(Predictions!J2:J5000,"{conf}",Predictions!K2:K5000,"<>"),"-")',
        ).font = _font()
        ws.cell(i, 2).number_format = "0.0%"
        ws.cell(i, 2).alignment = _center()


def _ensure_result_formulas(ws, row_num: int) -> None:
    m_col = get_column_letter(COL_ACT_WINNER)
    b_col = get_column_letter(COL_AWAY)
    c_col = get_column_letter(COL_HOME)
    k_col = get_column_letter(COL_AWAY_ACT)
    l_col = get_column_letter(COL_HOME_ACT)
    f_col = get_column_letter(COL_PRED_WINNER)
    ws.cell(row_num, COL_ACT_WINNER).value = (
        f'=IF({k_col}{row_num}="","",IF({k_col}{row_num}>{l_col}{row_num},{b_col}{row_num},{c_col}{row_num}))'
    )
    ws.cell(row_num, COL_CORRECT).value = (
        f'=IF({m_col}{row_num}="","",IF({m_col}{row_num}={f_col}{row_num},"{CHECK_MARK}","{X_MARK}"))'
    )


def _game_pk_from_payload(game: dict[str, Any]) -> int | None:
    value = game.get("game", {}).get("gamePk") or game.get("gamePk")
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _pitcher_name(game: dict[str, Any], side: str) -> str:
    return str(game.get("pitchers", {}).get(side, {}).get("name") or "TBD")


def _prediction_identity(game: dict[str, Any], target: date) -> tuple[str, str, str, str, str]:
    return (
        str(target),
        str(game["away"]["full"]),
        str(game["home"]["full"]),
        _pitcher_name(game, "away"),
        _pitcher_name(game, "home"),
    )


def _existing_keys(ws) -> set[tuple[str, str, str, str, str]]:
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:
            keys.add(
                (
                    str(row[0]),
                    str(row[1]),
                    str(row[2]),
                    str(row[3] or ""),
                    str(row[4] or ""),
                )
            )
    return keys


def _existing_game_pks(ws) -> set[int]:
    game_pks: set[int] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < COL_GAME_PK:
            continue
        value = row[COL_GAME_PK - 1]
        if value is None:
            continue
        try:
            game_pks.add(int(value))
        except (TypeError, ValueError):
            continue
    return game_pks


def _static_prediction_maps(
    target: date,
) -> tuple[dict[tuple[str, str, str, str, str], int], dict[tuple[str, str], list[int]]]:
    static = Path("public") / "slates" / f"{target}.json"
    if not static.exists():
        return {}, {}

    data = json.loads(static.read_text(encoding="utf-8"))
    by_identity: dict[tuple[str, str, str, str, str], int] = {}
    by_matchup: dict[tuple[str, str], list[int]] = defaultdict(list)
    for game in data.get("games", []):
        game_pk = _game_pk_from_payload(game)
        if game_pk is not None:
            by_identity[_prediction_identity(game, target)] = game_pk
            by_matchup[(str(game["away"]["full"]), str(game["home"]["full"]))].append(game_pk)
    return by_identity, by_matchup


def _backfill_game_pks_from_static(ws) -> int:
    dates = sorted(
        {
            date.fromisoformat(str(row[0])[:10])
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0]
        }
    )
    maps = {target: _static_prediction_maps(target) for target in dates}
    filled = 0
    assigned: dict[date, set[int]] = defaultdict(set)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and len(row) >= COL_GAME_PK and row[COL_GAME_PK - 1] is not None:
            try:
                assigned[date.fromisoformat(str(row[0])[:10])].add(int(row[COL_GAME_PK - 1]))
            except (TypeError, ValueError):
                pass

    for row in ws.iter_rows(min_row=2):
        if row[COL_GAME_PK - 1].value is not None or not row[0].value:
            continue
        target = date.fromisoformat(str(row[0].value)[:10])
        by_identity, by_matchup = maps.get(target, ({}, {}))
        key = (
            str(row[0].value),
            str(row[1].value or ""),
            str(row[2].value or ""),
            str(row[3].value or ""),
            str(row[4].value or ""),
        )
        game_pk = by_identity.get(key)
        if game_pk is None:
            matchup_key = (str(row[1].value or ""), str(row[2].value or ""))
            game_pk = next(
                (candidate for candidate in by_matchup.get(matchup_key, []) if candidate not in assigned[target]),
                None,
            )
        if game_pk is not None:
            row[COL_GAME_PK - 1].value = game_pk
            assigned[target].add(game_pk)
            filled += 1
    return filled


def _apply_row_style(ws, row_num: int) -> None:
    bg = CARD if row_num % 2 == 0 else PARCH
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row_num, col)
        cell.fill = _fill(bg)
        cell.font = _font()
        cell.border = _border()
        if col in (COL_AWAY_PRED, COL_HOME_PRED, COL_WIN_PROB, COL_AWAY_ACT, COL_HOME_ACT):
            cell.alignment = _center()


def _apply_conditional_formatting(ws) -> None:
    green_fill = _fill("C6EFCE")
    red_fill = _fill("FFC7CE")
    green_font = Font(name="Courier New", color="276221", bold=True, size=10)
    red_font = Font(name="Courier New", color="9C0006", bold=True, size=10)
    col = get_column_letter(COL_CORRECT)
    rng = f"{col}2:{col}5000"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=[f'"{CHECK_MARK}"'], fill=green_fill, font=green_font),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=[f'"{X_MARK}"'], fill=red_fill, font=red_font),
    )


def _load_predictions(target: date) -> list[dict[str, Any]]:
    static = Path("public") / "slates" / f"{target}.json"
    if static.exists():
        data = json.loads(static.read_text(encoding="utf-8"))
        return data.get("games", [])

    from backend.services.assemble import build_slate_payloads

    print(f"No static slate found for {target}, running pipeline...")
    return build_slate_payloads(target)


def log_predictions(target: date) -> int:
    games = _load_predictions(target)
    if not games:
        print(f"No games found for {target}")
        return 0

    wb = _load_or_create()
    ws = wb["Predictions"]
    existing = _existing_keys(ws)
    existing_game_pks = _existing_game_pks(ws)
    added = 0

    for game in games:
        game_pk = _game_pk_from_payload(game)
        fallback_key = _prediction_identity(game, target)
        if (game_pk is not None and game_pk in existing_game_pks) or (
            game_pk is None and fallback_key in existing
        ):
            continue

        prediction = game["prediction"]
        prob = max(prediction["awayProb"], prediction["homeProb"])
        row_num = ws.max_row + 1

        ws.cell(row_num, 1, str(target))
        ws.cell(row_num, 2, game["away"]["full"])
        ws.cell(row_num, 3, game["home"]["full"])
        ws.cell(row_num, 4, _pitcher_name(game, "away"))
        ws.cell(row_num, 5, _pitcher_name(game, "home"))
        ws.cell(row_num, 6, prediction["winner"])
        ws.cell(row_num, 7, prediction["awayRuns"])
        ws.cell(row_num, 8, prediction["homeRuns"])
        ws.cell(row_num, 9, prob / 100)
        ws.cell(row_num, 10, prediction["confLabel"])
        ws.cell(row_num, COL_GAME_PK, game_pk)
        _ensure_result_formulas(ws, row_num)

        ws.cell(row_num, 9).number_format = "0%"
        _apply_row_style(ws, row_num)
        added += 1
        existing.add(fallback_key)
        if game_pk is not None:
            existing_game_pks.add(game_pk)

    _apply_conditional_formatting(ws)
    wb.save(TRACKER_PATH)
    print(f"Logged {added} games for {target} -> {TRACKER_PATH}")
    return added


def _scores_for_date(target: date) -> tuple[dict[int, tuple[int, int]], dict[tuple[str, str], list[tuple[int, int]]]]:
    response = requests.get(
        "https://statsapi.mlb.com/api/v1/schedule",
        params={"date": str(target), "sportId": 1, "gameType": "R"},
        timeout=15,
    )
    response.raise_for_status()
    data = response.json()

    by_game_pk: dict[int, tuple[int, int]] = {}
    by_matchup: dict[tuple[str, str], list[tuple[int, int]]] = defaultdict(list)
    for date_rec in data.get("dates", []):
        for game in date_rec.get("games", []):
            if game.get("status", {}).get("abstractGameState") != "Final":
                continue

            teams = game.get("teams", {})
            away = teams.get("away", {})
            home = teams.get("home", {})
            away_name = away.get("team", {}).get("name", "")
            home_name = home.get("team", {}).get("name", "")
            away_score = away.get("score")
            home_score = home.get("score")
            if not away_name or not home_name or away_score is None or home_score is None:
                continue

            score = (int(away_score), int(home_score))
            try:
                by_game_pk[int(game["gamePk"])] = score
            except (KeyError, TypeError, ValueError):
                pass
            by_matchup[(away_name, home_name)].append(score)

    return by_game_pk, by_matchup


def _pending_actual_dates(ws) -> list[date]:
    dates: set[date] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if row[COL_AWAY_ACT - 1] is not None and row[COL_HOME_ACT - 1] is not None:
            continue
        try:
            dates.add(date.fromisoformat(str(row[0])[:10]))
        except ValueError:
            continue
    return sorted(dates)


def _fill_actuals_in_workbook(wb, target: date) -> int:
    ws = wb["Predictions"]
    by_game_pk, by_matchup = _scores_for_date(target)
    if not by_game_pk and not by_matchup:
        print(f"No completed games found for {target} (games may still be in progress)")
        return 0

    consumed_matchups: dict[tuple[str, str], int] = defaultdict(int)
    filled = 0
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) != str(target):
            continue
        if row[COL_AWAY_ACT - 1].value is not None and row[COL_HOME_ACT - 1].value is not None:
            continue

        score = None
        game_pk_value = row[COL_GAME_PK - 1].value
        if game_pk_value is not None:
            try:
                score = by_game_pk.get(int(game_pk_value))
            except (TypeError, ValueError):
                score = None

        if score is None:
            key = (str(row[COL_AWAY - 1].value or ""), str(row[COL_HOME - 1].value or ""))
            idx = consumed_matchups[key]
            if idx < len(by_matchup.get(key, [])):
                score = by_matchup[key][idx]
                consumed_matchups[key] += 1

        if score is None:
            continue

        row[COL_AWAY_ACT - 1].value = score[0]
        row[COL_HOME_ACT - 1].value = score[1]
        row[COL_AWAY_ACT - 1].alignment = _center()
        row[COL_HOME_ACT - 1].alignment = _center()
        _ensure_result_formulas(ws, row[0].row)
        filled += 1

    return filled


def fill_actuals(target: date) -> int:
    wb = _load_or_create()
    filled = _fill_actuals_in_workbook(wb, target)
    wb.save(TRACKER_PATH)
    print(f"Filled actuals for {filled} games on {target} -> {TRACKER_PATH}")
    return filled


def fill_missing_actuals() -> int:
    wb = _load_or_create()
    ws = wb["Predictions"]
    total = 0
    for target in _pending_actual_dates(ws):
        total += _fill_actuals_in_workbook(wb, target)
    wb.save(TRACKER_PATH)
    print(f"Filled actuals for {total} pending games -> {TRACKER_PATH}")
    return total


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--log", action="store_true", help="Append today's predictions")
    parser.add_argument("--actuals", action="store_true", help="Fill in actual scores")
    parser.add_argument("--all-missing", action="store_true", help="Fill actuals for all pending dates")
    parser.add_argument("--date", default=None, help="ISO date (default: today / yesterday for actuals)")
    args = parser.parse_args()

    configure_logging()

    if not args.log and not args.actuals:
        parser.print_help()
        return

    today = date.fromisoformat(args.date) if args.date else today_in_schedule_timezone()

    if args.log:
        log_predictions(today)

    if args.actuals:
        if args.all_missing:
            fill_missing_actuals()
        else:
            actuals_date = today if args.date else today - timedelta(days=1)
            fill_actuals(actuals_date)


if __name__ == "__main__":
    main()
